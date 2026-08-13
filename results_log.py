"""
results_log.py — append test measurements to an Excel (.xlsx) workbook.

Each CAPTURE (one Acquire or one I2C run) is written as a group of rows — one
row per measurement — sharing a single MERGED capture-number cell and a single
merged timestamp cell. New files get a header row; existing files are appended
to (load-modify-save), so the file must not be open in Excel during a run.
"""

import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

RESULTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "results")

HEADERS = ["Capture", "Measurement", "Value", "Units",
           "Limit min (ns)", "Limit max (ns)", "Status", "Note", "Timestamp"]
_CAPTURE_COL = 1
_TIMESTAMP_COL = len(HEADERS)   # last column


def resolve_log_path(filename: str) -> str:
    """
    Resolve a user-entered log name to an .xlsx path under results/.

    Any extension is normalized to .xlsx (merged cells need Excel, not CSV).
    Only the basename is used so paths cannot escape the results folder.
    """
    name = filename.strip()
    if not name:
        raise ValueError("Enter a log file name (e.g. measurements.xlsx).")
    base = os.path.splitext(os.path.basename(name))[0]
    os.makedirs(RESULTS_DIR, exist_ok=True)
    return os.path.join(RESULTS_DIR, base + ".xlsx")


def _cell(x):
    """Numbers pass through (Excel stores them numeric); None -> blank."""
    if x is None:
        return ""
    if isinstance(x, float):
        return float(f"{x:.6g}")
    return x


class ResultLogger:
    def __init__(self, path: str):
        self.path = path
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        self.run_id = datetime.now().strftime("%Y%m%d-%H%M%S")

    def log(self, measurement, value, *, units="", status="",
            limit_min_ns=None, limit_max_ns=None, note=""):
        """Log a single measurement as a one-row capture."""
        return self.log_capture([{
            "measurement": measurement, "value": value, "units": units,
            "status": status, "limit_min_ns": limit_min_ns,
            "limit_max_ns": limit_max_ns, "note": note,
        }])

    def log_capture(self, rows):
        """
        Append one capture (list of measurement dicts) as a group of rows.

        The capture number and timestamp each go in ONE cell merged down the
        group's rows; each measurement gets its own row. A single-measurement
        capture is one row (nothing to merge).
        """
        rows = [r for r in rows if r]
        if not rows:
            return None
        ts = datetime.now().isoformat(timespec="seconds")

        if os.path.exists(self.path) and os.path.getsize(self.path) > 0:
            wb = load_workbook(self.path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "measurements"
            ws.append(HEADERS)

        capture = self._next_capture(ws)
        start = ws.max_row + 1
        for r in rows:
            ws.append([
                capture,
                r.get("measurement", ""),
                _cell(r.get("value")),
                r.get("units", ""),
                _cell(r.get("limit_min_ns")),
                _cell(r.get("limit_max_ns")),
                r.get("status", ""),
                r.get("note", ""),
                ts,
            ])
        end = ws.max_row
        if end > start:   # merge the capture number + timestamp across the group
            for col in (_CAPTURE_COL, _TIMESTAMP_COL):
                ws.merge_cells(start_row=start, start_column=col,
                               end_row=end, end_column=col)
                ws.cell(row=start, column=col).alignment = Alignment(
                    vertical="center")

        try:
            wb.save(self.path)
        except PermissionError as e:
            raise PermissionError(
                f"Could not write {self.path} — is it open in Excel? "
                "Close it and run again."
            ) from e
        return capture

    @staticmethod
    def _next_capture(ws) -> int:
        """Next capture number = max existing (top cell of each merge) + 1."""
        max_cap = 0
        for (v,) in ws.iter_rows(min_row=2, min_col=_CAPTURE_COL,
                                 max_col=_CAPTURE_COL, values_only=True):
            if isinstance(v, int):
                max_cap = max(max_cap, v)
            elif isinstance(v, str) and v.isdigit():
                max_cap = max(max_cap, int(v))
        return max_cap + 1
